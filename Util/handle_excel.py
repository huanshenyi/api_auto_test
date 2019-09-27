import openpyxl
import sys
import os
base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)


open_excel = openpyxl.load_workbook(base_path + "/Case/api_test_case.xlsx")
sheet_name = open_excel.sheetnames
excel_value = open_excel[sheet_name[0]]


class HandExcel:
    def load_excel(self):
        """
        excelロード
        :return:
        """
        open_excel = openpyxl.load_workbook(base_path + "/Case/api_test_case_imooc.xlsx")
        return open_excel

    def get_sheet_data(self, index=None):
        """
        シートロード ,どのシート読み込むのかも指定可能
        :return:
        """
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, column):
        """
        シートのカラム内容取得
        :return:
        """
        data = self.get_sheet_data().cell(row=row, column=column).value
        return data

    def get_rows(self):
        """
        行数を取得
        :return:
        """
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        """
        行のvalueを取得
        :param row:
        :return:
        """
        row_list = []
        # 1行の内容
        for i in self.get_sheet_data()[row]:
             row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        """
        データーの入力
        :param row:
        :param cols:
        :param value:
        :return:
        """
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(base_path + "/Case/api_test_case_imooc.xlsx")



excel_data = HandExcel()

if __name__ == "__main__":
    handel = HandExcel()
    handel.excel_write_data(2, 11, "成功")
